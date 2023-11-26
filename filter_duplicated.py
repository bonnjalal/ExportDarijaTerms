from openpyxl import Workbook, load_workbook


wb1 = load_workbook('Darija_Terms_EN.xlsx')
ws1 = wb1.active  

wb2 = Workbook()  
ws2 = wb2.active

values = []
for row in ws1.iter_rows(min_row=5, max_row=8774):
    if row[0].value in values:
        pass
    else:
        values.append(row[0].value)
        ws2.append((cell.value for cell in row[0:99]))

# for value in values:
#   ws2.append([value])










wb2.save('Filtered_Darija_Terms_EN.xlsx')


