from queue import Empty
import re
import json
# import xlwings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment  
import openpyxl
import copy

with open('test.json') as json_file:
    data = json.load(json_file)
     
        # Print the type of data variable
        # print(data)

wb = Workbook()  
ws = wb.active
# wb1.save('test.xlsx')
# wb = load_workbook('test.xlsx', data_only=True)
# ws = wb.active 

def depth(d):
    if isinstance(d, dict):
        return 1 + (max(map(depth, d.values())) if d else 0)
    return 0


def getMergedCellVal(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value

def getValueWithMergeLookup(cell):
    idx = cell.coordinate
    for range_ in ws.merged_cells.ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(str(range_)))
        for row in merged_cells:
            if idx in row:
                value = ws[merged_cells[0][0]].value
                return value

    # return sheet.cell(idx).value
    return ws[idx].value

def wordfinder2(searchString:str, rngr:list, rngc:list):
    rowrng = range(rngr[0], rngr[1]+1)  
    colrng = range(rngc[0], rngc[1]+1)
    print("serching for: " + searchString)
    for i in rowrng:
        for j in colrng:
            cell = ws.cell(row=i,column=j)
            strValue = cell.value
            if type(cell) == openpyxl.cell.cell.MergedCell:
                # coord = cell.coordinate
                strValue = getValueWithMergeLookup(cell)
                print(strValue)

            if searchString == strValue:
                print("found found found found found ")
                return i,j
    return -1,-1

def wordfinder(searchString:str, rngr:list, rngc:list):
    # print("serching for: " + searchString)
    for row in ws.iter_cols(min_row=rngr[0], min_col=rngc[0], max_row=rngr[1], max_col=rngc[1]):  
        for cell in row:  
            strValue = cell.value 
            if type(cell) == openpyxl.cell.cell.MergedCell:
                coord = cell.coordinate
                strValue = getValueWithMergeLookup(cell)
                # print("strVal: " + str(strValue))
            if searchString == strValue:
                print("found found found found found ")
                return cell.row,cell.column
    return -1,-1

def getMergedCellRange(cell):
    rng = [s for s in ws.merged_cells.ranges if cell.coordinate in s]
    if len(rng) != 0:
        return [rng[0].min_row, rng[0].max_row], [rng[0].min_col, rng[0].max_col]

    return [cell.row, cell.row], [cell.column, cell.column]

def cellEmpty(cell):
    value = cell.value
    # if type(cell) == openpyxl.cell.cell.MergedCell:
    #     return False
    if value is None or value == "":
        return True
    return False

rangeList = [[[1,1],[2,2]]]
# rangeDict = {}
countEnter = 0
countExit = 0
count = [0,0]

def nested_get(dic, keys):
    dic2 = copy.deepcopy(dic)
    for key in keys:
        dic2 = dic2[key]
    return dic2

def nested_set(dic, keys, value):
    for key in keys[:-1]:
        dic = dic.setdefault(key, {})
    dic[keys[-1]] = value

def nested_del(dic, keys):
    for key in keys[:-1]:
        dic = dic[key]
    del dic[keys[-1]]

def getAllDicKeys(d):
    keyList = []
    for key, value in d.items():
        keyList.append(key)
        if isinstance(value, dict):
            get_all_keys(value)
    return keyList

def updateDic(dic,diffrence):

    data = copy.deepcopy(dic)
    for k, v in data.copy().items():
        
        if isinstance(v, dict):     # For DICT
            data[k] = updateDic(v,diffrence)
        elif isinstance(v, list):   # For LIST
            # diffrence = col1 - diffrence
            tmpRng = [v[0], [v[1][0]+diffrence ,v[1][1]+diffrence]]
            data[k] = tmpRng
        # elif v == 'CHANGE ME':      # Update Key-Value
        #     # data.pop(k)
        #     # OR
        #     del data[k]
        #     data[f"{k}.CHANGED"] = 'CHANGED'
    # print(rangesDict)
    return data

rangesDict = {"word": {"rng" : [[1,1],[2,2]]}}
keysList = ["word"]
def insertCell(data, crow, depth1):
    
    d = 0
    dp = depth1
    insertCol = True
    for it2 in data:
        tmpKeysList1 = []
        for i in range(dp):
            tmpKeysList1.append(keysList[i])
        rng = nested_get(rangesDict, tmpKeysList1+["rng"]) 
        row2,col1 = wordfinder(it2, [dp,dp], rng[1])
        print(tmpKeysList1)
        print("search range col: " + str(rng[1]) + " dp " + str(dp))
        col2 = col1
        
        if row2 == -1:
            col1 = rng[1][0]
            maxCol = rng[1][1]
            row2 = dp
            cCell = ws.cell(row=row2, column=maxCol)

            if not cellEmpty(cCell) or not insertCol:
                maxCol += 1
                ws.insert_cols(idx=maxCol)
                col1 = maxCol
            col2 = maxCol

            keysListTmp = [] 
            for r in range(dp):
                if r == 0:
                    keysListTmp.append(keysList[r])
                    tmpRng = nested_get(rangesDict, keysListTmp + ["rng"])
                    tmpRng[1][1] = ws.max_column + 1
                    nested_set(rangesDict, keysListTmp + ["rng"], tmpRng) 
                else:
                    keysListTmp.append(keysList[r])
                    tmpRng = nested_get(rangesDict, keysListTmp + ["rng"])
                    tmpRng[1][1]= col2
                    nested_set(rangesDict, keysListTmp + ["rng"], tmpRng) 


            ws.cell(row=row2, column=col2).value = it2
        
            rngr = [dp, dp]
            rngc = [col1, col2]
            
            tmpKeysList1.append(it2)
            tmpKeysList1.append("rng")
            nested_set(rangesDict, tmpKeysList1, [rngr, rngc])
        else:

            
            if dp == 1:
                it2rng = rangesDict['word'][it2]['rng']
                diffrence = col1 - it2rng[1][0]
                newData3 = updateDic(rangesDict["word"][it2], diffrence)
                rangesDict["word"][it2] = newData3
                print(rangesDict)

            # it2rng = nested_get(rangesDict, tmpKeysList1+[it2,"rng"])
            # diffrence = col1 - it2rng[1][0]
            # tmpRng = [it2rng[0], [it2rng[1][0]+diffrence ,it2rng[1][1]+diffrence]]
            #
            # nested_set(rangesDict, tmpKeysList1+[it2,"rng"], tmpRng)
            
            # rangesDict["word"]["rng"] = [[1,1],[2,ws.max_column]]
        try: 
            keysList[dp] = it2
        except:
            keysList.append(it2)
            
        if depth(data[it2]) <= d:
            value = data[it2]
            if (type(value) == list):
                value = ', '.join([str(elem) for elem in value])
            ws.cell(row=crow, column=col1).value = value
        else:
            insertCell(data[it2], crow, dp+1)

        insertCol = False

def get_all_keys(d):
    for key, value in d.items():
        yield key
        if isinstance(value, dict):
            yield from get_all_keys(value)

def getAllRngValues(d):
    for key, value in d.items():
        if key == "rng":
            yield value
        if isinstance(value, dict):
            yield from getAllRngValues(value)

def mergeRanges():
    i = 0
    for x in getAllRngValues(rangesDict): 
        if i != 0:
            print("merging ....")
            print(x)
            ws.merge_cells(start_row=x[0][0], start_column=x[1][0], end_row=x[0][1] , end_column=x[1][1] )
            cell = ws.cell(row=x[0][0], column=x[1][0])  
            # # cell.value = 'Devansh Sharma'  
            cell.alignment = Alignment(horizontal='center', vertical='center')
        i+=1

def insertData():
    crow = 5

    for it in data.keys():
        crow+=1
        ws.cell(row=crow, column=1).value = it
        insertCell(data[it], crow, 1)

    mergeRanges()

insertData()
wb.save("test.xlsx")  


