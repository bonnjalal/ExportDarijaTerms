import requests
from openpyxl import Workbook, load_workbook

wb2 = Workbook()  
ws2 = wb2.active


def query(request):
    request['action'] = 'query'
    request['format'] = 'json'
    lastContinue = {}
    while True:
        # Clone original request
        req = request.copy()
        # Modify it with the values returned in the 'continue' section of the last result.
        req.update(lastContinue)
        URL = "https://incubator.wikimedia.org/w/api.php"
        # Call API
        result = requests.get(URL, params=req).json()
        if 'error' in result:
            raise Exception(result['error'])
        if 'warnings' in result:
            print(result['warnings'])
        if 'query' in result:
            yield result['query']
        if 'continue' not in result:
            break
        lastContinue = result['continue']

PARAMS = {
    "action": "query",
    "format": "json",
    "list": "allpages",
    "apfrom": "Wt/ary/",
    "aplimit":500,
}
crow = 1
for result in query(PARAMS):
    # print(result)
    PAGES = result["allpages"]

    # print(DATA)
    for page in PAGES:
        title = page["title"]
        print(title)
        word = title.replace('Wt/ary/', '')
        ws2.cell(row= crow, column=1).value = word
        crow += 1
        if "Wt/ary" not in title:
            break
    else:
        continue
    break
    

wb2.save('Darija_Terms_ary.xlsx')
