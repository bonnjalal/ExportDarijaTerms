#!/usr/bin/env python
import os
import pywikibot
import requests
import time
import csv
# import json
from operator import itemgetter
from extract_init import ExtractHelper
from dicToXlsx import DicToExcel
from openpyxl import Workbook
# from openpyxl.styles import Alignment  
# import openpyxl


extH = ExtractHelper()

site = pywikibot.Site("en", "wiktionary")

#user = pywikibot.User(site, "Bonnjalal00")
def getSecInfo(secDict, line):
    for x in secDict["parse"]["sections"]:
        if x["line"] == line:
            return x["index"], x["number"]
    return "-1.0","-1.0"

def isEty1InSec(text, name):
    if name in text:
        return True
    return False

def getAllSect(secDict, startNum):
    startNum += "."
    sections = {}
    for x in secDict["parse"]["sections"]:
        num = x["number"]
        if num.startswith(startNum):
            sections[num] = x["index"]
    return sections

def isArticle(page):
    # print(page.namespace().normalize_name())
    if str(page.namespace()) == ":":
        return True
    return False
    # names = ["Module","Template","User","Wiktionary","File","MediaWiki","Thread","Help","Summary","Appendix",
    #          "Concordance","Rhymes","Transwiki","Thesaurus","Citations","Sign gloss","Reconstruction", ]
    # for name in names:
    #     if title.startswith(name+":"):
    #         return false 
    # return true



wb = Workbook()  
ws = wb.active
dToE = DicToExcel(wb, ws)
crow = 5

def makeNetworkReq(req):
    try:
        return requests.get(req)
    except:
        print("sleep .... 30 sec")
        time.sleep(30)
        return requests.get(req)

def getArticlesData(pagesList):

    # fp =  open('test.json', 'w')
    sectNames = ["Etymology", "Pronunciation", "Verb", "Noun", "Adverb", "Adjective", "Pronoun", "Synonyms", "Alternative forms"]
    for page in pagesList:
        title = page.title()
        print(title)
        if isArticle(page):
            # print(title)
            # sectionsJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=sections&page={}&format=json".format(title))
            sectionsJson = makeNetworkReq("https://en.wiktionary.org/w/api.php?action=parse&prop=sections&page={}&format=json".format(title))
            sectionsDict = sectionsJson.json()
            secIndex, secNum = getSecInfo(sectionsDict,"Moroccan Arabic") 
            
            # print("isEty1: " + str(isEty1))

            if secNum != "-1.0":
                allSections = getAllSect(sectionsDict, secNum)
                # print("secIndex " + secIndex + " / secNumber " + secNum)
                # mainSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(secIndex,title))
                mainSectJson = makeNetworkReq("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(secIndex,title))
                mainSectionText = mainSectJson.json()
                # print(mainSectionText["parse"]["wikitext"]["*"])
                isEty1 = isEty1InSec(mainSectionText["parse"]["wikitext"]["*"], "Etymology 1") 
                isEty = isEty1InSec(mainSectionText["parse"]["wikitext"]["*"], "=Etymology=")

                sectionDic = {}
                # [title, depth, dictionary]
                dicList = []

                depth2title = ""
                depth3title = ""
                depth4title = ""
 
                if isEty1:
                    secDepth = 2
                    sortN = 1
                    for num, index in allSections.items():
                        # print(num,index)
                        # selectedSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                        selectedSectJson = makeNetworkReq("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                        sectionText = selectedSectJson.json()
                        secTxt = sectionText["parse"]["wikitext"]["*"]
                        secTitle = secTxt.splitlines()[0].replace('=', '')
                        # secDepth = len(num.strip().replace('.', ''))
                        # secDepth = num.count('.') + 1 
                        
                        
                        # print(secTitle)
                        # print(secTxt)
                        # print(secTitle)
                        if "Etymology 1" == secTitle:
                            secDepth = 2
                            sortN = 1
                        elif secTitle == "Etymology 2":
                            secDepth = 2
                            sortN = 2
                        elif "Etymology" in secTitle:
                            break
                        else:
                            secDepth = 3
                        # depth5title = ""

                        for sectName in sectNames: 
                            if sectName in secTitle:
                                secN = sectName
                                if secDepth == 2:
                                    secN = secTitle
                                sectDic = extH.getSectContent(sectName, secTxt)
                                sectList = [secN,sortN, secDepth, sectDic]
                                dicList.append(sectList)
                                break

                    sortedDicList = sorted(dicList, key=itemgetter(1,2)) 
                    # print(sortedDicList) 
                    # dicList = []
                    for dic in sortedDicList:
                        secDepth = dic[2]
                        secTitle = dic[0]
                        sectDic = dic[3]
                        if secDepth == 2:
                            depth2title = secTitle
                            sectionDic[secTitle] = sectDic
                        elif secDepth == 3: 
                            depth3title = secTitle
                            sectionDic[depth2title][secTitle] = sectDic
                        elif secDepth == 4:
                            depth4title = secTitle
                            sectionDic[depth2title][depth3title][secTitle] = sectDic
                        elif secDepth == 5:
                            sectionDic[depth2title][depth3title][depth4title][secTitle] = sectDic

                elif not isEty:
                    # secDepth = 2
                    sectList = ["Etymology 1", 2, {}]
                    dicList.append(sectList)
                    secDepth = 3

                    for num, index in allSections.items():
                        # print(num,index)
                        # selectedSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                        selectedSectJson = makeNetworkReq("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                        sectionText = selectedSectJson.json()
                        secTxt = sectionText["parse"]["wikitext"]["*"]
                        secTitle = secTxt.splitlines()[0].replace('=', '')
                        # secDepth = len(num.strip().replace('.', ''))
                        # secDepth = num.count('.') + 1 
                        
                        
                        # if secTitle not in sectNames[0]:
                            # secDepth = 3

                        # secDepth+=1

                        for sectName in sectNames: 
                            if sectName in secTitle:
                                secN = sectName
                                if secDepth == 2:
                                    secN = secTitle
                                sectDic = extH.getSectContent(sectName, secTxt)
                                sectList = [secN, secDepth, sectDic]
                                dicList.append(sectList)
                                break

                    sortedDicList = sorted(dicList, key=itemgetter(1)) 
                    # print(sortedDicList) 
                    # dicList = []
                    for dic in sortedDicList:
                        secDepth = dic[1]
                        secTitle = dic[0]
                        sectDic = dic[2]
                        if secDepth == 2:
                            depth2title = secTitle
                            sectionDic[secTitle] = sectDic
                        elif secDepth == 3: 
                            depth3title = secTitle
                            sectionDic[depth2title][secTitle] = sectDic
                        elif secDepth == 4:
                            depth4title = secTitle
                            sectionDic[depth2title][depth3title][secTitle] = sectDic
                        elif secDepth == 5:
                            sectionDic[depth2title][depth3title][depth4title][secTitle] = sectDic

                elif not isEty1:
                    secDepth = 2
                    for num, index in allSections.items():
                        # print(num,index)
                        # selectedSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                        selectedSectJson = makeNetworkReq("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                        sectionText = selectedSectJson.json()
                        secTxt = sectionText["parse"]["wikitext"]["*"]
                        secTitle = secTxt.splitlines()[0].replace('=', '')
                        # secDepth = len(num.strip().replace('.', ''))
                        # secDepth = num.count('.') + 1 
                        
                        # print(secTitle)
                        # print(secTxt)
                        
                        # depth5title = ""

                          
                        if sectNames[0] in secTitle:
                            secTitle = "Etymology 1"
                            secDepth = 2
                        else:
                            secDepth = 3

                        for sectName in sectNames: 
                            if sectName in secTitle:
                                secN = sectName
                                if secDepth == 2:
                                    secN = secTitle
                                sectDic = extH.getSectContent(sectName, secTxt)
                                sectList = [secN, secDepth, sectDic]
                                dicList.append(sectList)
                                break

                    sortedDicList = sorted(dicList, key=itemgetter(1)) 
                    # print(sortedDicList) 
                    # dicList = []
                    for dic in sortedDicList:
                        secDepth = dic[1]
                        secTitle = dic[0]
                        sectDic = dic[2]
                        if secDepth == 2:
                            depth2title = secTitle
                            sectionDic[secTitle] = sectDic
                        elif secDepth == 3: 
                            depth3title = secTitle
                            sectionDic[depth2title][secTitle] = sectDic
                        elif secDepth == 4:
                            depth4title = secTitle
                            sectionDic[depth2title][depth3title][secTitle] = sectDic
                        elif secDepth == 5:
                            sectionDic[depth2title][depth3title][depth4title][secTitle] = sectDic



                if (len(sectionDic) > 2):
                    # print("Iam here above 2")
                    wordDict = {title: {"Etymology 1":sectionDic["Etymology 1"], "Etymology 2":sectionDic["Etymology 2"]}}
                else: 
                    wordDict = {title:sectionDic}
                # print(wordDict)
                global crow
                crow+=1
                ws.cell(row=crow, column=1).value = title
                dToE.insertCell(wordDict[title], crow, 1)

                # print(wordDict)
                # json.dump(wordDict, fp, indent=4)
                # fp.write(',\n')
                    # json.dump(wordDict, fp, indent=4)

                


                    # print(secTitle)
                    # print(num.count("."))
    # print(dToE.rangesDict)

def getArticlesData2(pagesList):

    # fp =  open('test.json', 'w')
    sectNames = ["Etymology", "Pronunciation", "Verb", "Noun", "Adverb", "Adjective", "Pronoun", "Synonyms", "Alternative forms"]
    for page in pagesList:
        title = page.title()
        print(title)
        if isArticle(page):
            # print(title)
            sectionsJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=sections&page={}&format=json".format(title))
            sectionsDict = sectionsJson.json()
            secIndex, secNum = getSecInfo(sectionsDict,"Moroccan Arabic") 
            allSections = getAllSect(sectionsDict, secNum)
            # print("secIndex " + secIndex + " / secNumber " + secNum)
            mainSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(secIndex,title))
            mainSectionText = mainSectJson.json()
            # print(mainSectionText["parse"]["wikitext"]["*"])
            isEty1 = isEty1InSec(mainSectionText["parse"]["wikitext"]["*"], "Etymology 1") 
            isEty = isEty1InSec(mainSectionText["parse"]["wikitext"]["*"], "=Etymology=") 
            # print("isEty1: " + str(isEty1))

            if secNum != "-1.0":

                sectionDic = {}
                # [title, depth, dictionary]
                dicList = []

                depth2title = ""
                depth3title = ""
                depth4title = ""
                # for num, index in allSections.items():
                #     # print(num,index)
                #     selectedSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                #     sectionText = selectedSectJson.json()
                #     secTxt = sectionText["parse"]["wikitext"]["*"]
                #     secTitle = secTxt.splitlines()[0].replace('=', '')
                #     # secDepth = len(num.strip().replace('.', ''))
                #     secDepth = num.count('.') + 1 
                #     
                #     # print(secTitle)
                #     # print(secTxt)
                #     
                #     # depth5title = ""
                #
                #     if not isEty1:
                #         # depth2title = "Etymology 1"
                #         # sectionDic[depth2title] = {}
                #        
                #         if secTitle == sectNames[0]:
                #             secTitle = "Etymology 1"
                #         else:
                #             secDepth +=1
                #
                #
                #     for sectName in sectNames:
                #         if isEty1 and sectName in secTitle:
                #             sectDic = extH.getSectContent(sectName, secTxt)
                #             # sectList = [secTitle, secDepth, sectDic]
                #             # dicList.append(sectList)
                #             if secDepth == 2:
                #                 depth2title = secTitle
                #                 sectionDic[secTitle] = sectDic
                #             elif secDepth == 3: 
                #                 depth3title = secTitle
                #                 sectionDic[depth2title][secTitle] = sectDic
                #             elif secDepth == 4:
                #                 depth4title = secTitle
                #                 sectionDic[depth2title][depth3title][secTitle] = sectDic
                #             elif secDepth == 5:
                #                 sectionDic[depth2title][depth3title][depth4title][secTitle] = sectDic
                #             break
                #         elif sectName in secTitle and not isEty1:
                #             sectDic = extH.getSectContent(sectName, secTxt)
                #             sectList = [secTitle, secDepth, sectDic]
                #             dicList.append(sectList)
                #
                # This will sort the outer list by items in the inner list (in this example by depth [index 2])
                # [title, depth, dictionary]
                # key=itemgetter(2,0,1) this will sort the list with multiplle items, starting from index 2

                # if not isEty1:
                #
                #     sortedDicList = sorted(dicList, key=itemgetter(1)) 
                #     print(sortedDicList) 
                #     # dicList = []
                #     for dic in sortedDicList:
                #         secDepth = dic[1]
                #         secTitle = dic[0]
                #         sectDic = dic[2]
                #         if secDepth == 2:
                #             depth2title = secTitle
                #             sectionDic[secTitle] = sectDic
                #         elif secDepth == 3: 
                #             depth3title = secTitle
                #             sectionDic[depth2title][secTitle] = sectDic
                #         elif secDepth == 4:
                #             depth4title = secTitle
                #             sectionDic[depth2title][depth3title][secTitle] = sectDic
                #         elif secDepth == 5:
                #             sectionDic[depth2title][depth3title][depth4title][secTitle] = sectDic
 
                if isEty1:
                    for num, index in allSections.items():
                        # print(num,index)
                        selectedSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                        sectionText = selectedSectJson.json()
                        secTxt = sectionText["parse"]["wikitext"]["*"]
                        secTitle = secTxt.splitlines()[0].replace('=', '')
                        # secDepth = len(num.strip().replace('.', ''))
                        secDepth = num.count('.') + 1 
                        
                        # print(secTitle)
                        # print(secTxt)
                        
                        # depth5title = ""

                        for sectName in sectNames:
                            if sectName in secTitle:
                                sectDic = extH.getSectContent(sectName, secTxt)
                                # sectList = [secTitle, secDepth, sectDic]
                                # dicList.append(sectList)
                                if secDepth == 2:
                                    depth2title = secTitle
                                    sectionDic[secTitle] = sectDic
                                elif secDepth == 3: 
                                    depth3title = secTitle
                                    sectionDic[depth2title][secTitle] = sectDic
                                elif secDepth == 4:
                                    depth4title = secTitle
                                    sectionDic[depth2title][depth3title][secTitle] = sectDic
                                elif secDepth == 5:
                                    sectionDic[depth2title][depth3title][depth4title][secTitle] = sectDic
                                break
                elif not isEty:
                    for num, index in allSections.items():
                        # print(num,index)
                        selectedSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                        sectionText = selectedSectJson.json()
                        secTxt = sectionText["parse"]["wikitext"]["*"]
                        secTitle = secTxt.splitlines()[0].replace('=', '')
                        # secDepth = len(num.strip().replace('.', ''))
                        secDepth = num.count('.') + 1 
                        
                        if secDepth == 2:
                            sectList = ["Etymology 1", secDepth, {}]
                            dicList.append(sectList)

                        secDepth+=1

                        for sectName in sectNames: 
                            if sectName in secTitle:
                                sectDic = extH.getSectContent(sectName, secTxt)
                                sectList = [secTitle, secDepth, sectDic]
                                dicList.append(sectList)
                                break

                    sortedDicList = sorted(dicList, key=itemgetter(1)) 
                    # print(sortedDicList) 
                    # dicList = []
                    for dic in sortedDicList:
                        secDepth = dic[1]
                        secTitle = dic[0]
                        sectDic = dic[2]
                        if secDepth == 2:
                            depth2title = secTitle
                            sectionDic[secTitle] = sectDic
                        elif secDepth == 3: 
                            depth3title = secTitle
                            sectionDic[depth2title][secTitle] = sectDic
                        elif secDepth == 4:
                            depth4title = secTitle
                            sectionDic[depth2title][depth3title][secTitle] = sectDic
                        elif secDepth == 5:
                            sectionDic[depth2title][depth3title][depth4title][secTitle] = sectDic

                elif not isEty1:
                    for num, index in allSections.items():
                        # print(num,index)
                        selectedSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
                        sectionText = selectedSectJson.json()
                        secTxt = sectionText["parse"]["wikitext"]["*"]
                        secTitle = secTxt.splitlines()[0].replace('=', '')
                        # secDepth = len(num.strip().replace('.', ''))
                        secDepth = num.count('.') + 1 
                        
                        # print(secTitle)
                        # print(secTxt)
                        
                        # depth5title = ""

                          
                        if secTitle == sectNames[0]:
                            secTitle = "Etymology 1"
                        else:
                            secDepth +=1

                        for sectName in sectNames: 
                            if sectName in secTitle:
                                sectDic = extH.getSectContent(sectName, secTxt)
                                sectList = [secTitle, secDepth, sectDic]
                                dicList.append(sectList)
                                break

                    sortedDicList = sorted(dicList, key=itemgetter(1)) 
                    # print(sortedDicList) 
                    # dicList = []
                    for dic in sortedDicList:
                        secDepth = dic[1]
                        secTitle = dic[0]
                        sectDic = dic[2]
                        if secDepth == 2:
                            depth2title = secTitle
                            sectionDic[secTitle] = sectDic
                        elif secDepth == 3: 
                            depth3title = secTitle
                            sectionDic[depth2title][secTitle] = sectDic
                        elif secDepth == 4:
                            depth4title = secTitle
                            sectionDic[depth2title][depth3title][secTitle] = sectDic
                        elif secDepth == 5:
                            sectionDic[depth2title][depth3title][depth4title][secTitle] = sectDic



                if (len(sectionDic) > 2):
                    wordDict = {title: {"Etymology 1":sectionDic["Etymology 1"], "Etymology 2":sectionDic["Etymology 2"]}}
                else: 
                    wordDict = {title:sectionDic}
                print(wordDict)
                global crow
                crow+=1
                ws.cell(row=crow, column=1).value = title
                dToE.insertCell(wordDict[title], crow, 1)

                # print(wordDict)
                # json.dump(wordDict, fp, indent=4)
                # fp.write(',\n')
                    # json.dump(wordDict, fp, indent=4)

                


                    # print(secTitle)
                    # print(num.count("."))
    # print(dToE.rangesDict)

def extractCatName(cat):
    title = cat
    return title.replace("Category:", "")

def main (catName = "Moroccan Arabic language"):
    cat = pywikibot.Category(site, catName)
    catInfo = cat.categoryinfo
    pagesCount = catInfo["pages"]
    subcatsCount = catInfo["subcats"]
    print(pagesCount)
    if  pagesCount >= 1:
        getArticlesData(list(cat.articles()))

    if subcatsCount >=1:
        catList = list(cat.subcategories())
        for category in catList:
            name = extractCatName(category.title())
            main(name)

# main("Moroccan Arabic interjections")
# main("Moroccan Arabic lemmas")
main("Moroccan Arabic non-lemma forms")
#Moroccan Arabic prepositions

# main()
dToE.saveExcel("Moroccan_Arabic_non-lemma_forms.xlsx")


with open('CategoryList.csv') as inf, open('CategoryList_tmp.csv', 'w') as outf:
    reader = csv.reader(inf)
    writer = csv.writer(outf)
    next(reader) # skip the header row
    csvList = ['Title', 'ArticlesCount','isExtracted']
    writer.writerow(csvList)
    for line in reader:
        if int(line[1]) >= 10 and line[2] != 'True':
            main(line[0])
            writer.writerow([line[0], line[1], 'True'])
        else:
            writer.writerow(line)

os.remove('CategoryList.csv')
os.rename('CategoryList_tmp.csv', 'CategoryList.csv')
# myreg()





# site = pywikibot.Site("en", "wiktio nary")
#user = pywikibot.User(site, "Bonnjalal00")
#cat = pywikibot.Category(site, "Moroccan Arabic templates")
# cat = pywikibot.Category(site, "Moroccan Arabic language")
#mem = cat.members()

#for user in range(10):
#    print (list(allusers).get(user))

# catInfo = cat.categoryinfo
#
# pagesCount = catInfo["pages"]
# subcatsCount = catInfo["subcats"]


#print(type(pagesCount))
#print(subcatsCount)

# print(pagesCount)
# if pagesCount >= 1:
#     getArticlesData(list(cat.articles()))
#
# catList = list(cat.subcategories())
#print(catList)
#title = list(catList[0].articles())[0].title()
# print(title)
#
# sectionsJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=sections&page={}&format=json".format(title))
# sectionsDict = sectionsJson.json()
# secIndex, secNum = getSecInfo(sectionsDict,"Moroccan Arabic") 
#
# allSections = getAllSect(sectionsDict, secNum)
#selectedSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(secIndex,title))
#sectionText = selectedSectJson.json()

## count char in a string ##
#string.count('.')

# text = list(catList[0].articles())[0].text
# print(sectionsDict["parse"]["sections"][17])

# for num, index in allSections.items():
#     print(num,index)
#     selectedSectJson = requests.get("https://en.wiktionary.org/w/api.php?action=parse&prop=wikitext&section={}&page={}&format=json".format(index,title))
#     sectionText = selectedSectJson.json()
#     print(sectionText["parse"]["wikitext"])
#     print(num.count("."))
# print(allSections)
#

## convert json to xlxs (needs to convert dict to json first)
# import  jpype     
#   import  asposecells     
#   jpype.startJVM() 
#   from asposecells.api import Workbook
#   workbook = Workbook("testData.json")
#   workbook.save("testData.xlsx")
#   jpype.shutdownJVM()

