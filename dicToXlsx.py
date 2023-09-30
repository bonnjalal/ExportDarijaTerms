import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment  
import openpyxl
import copy

with open('test.json') as json_file:
    data = json.load(json_file)

class DicToExcel:
    # wb = Workbook()  
    # ws = wb.active

    keysDict = {
        "Etymology 1": {
            "rng": [[1,1],[2,50]],
            "From 1": {
                "rng": [[2,2],[2,4]],
                "Language": {
                    "rng": [[3,3],[2,2]]
                },
                "Word": {
                    "rng": [[3,3],[3,3]]
                },
                "translation": {
                    "rng": [[3,3],[4,4]]
                }
            },
            "From 2": {
                "rng": [[2,2],[5,7]],
                "Language": {
                    "rng": [[3,3],[5,5]]
                },
                "Word": {
                    "rng": [[3,3],[6,6]]
                },
                "translation": {
                    "rng": [[3,3],[7,7]]
                }
            },
            "Pronunciation":{
                "rng": [[2,2],[8,12]],
                "pron_text 1": {
                    "rng": [[3,3],[8,9]],
                    "accent": {
                        "rng": [[4,4],[8,8]]
                    },
                    "text": {
                        "rng": [[4,4],[9,9]]
                    }

                },
                "pron_text 2": {
                    "rng": [[3,3],[10,11]],
                    "accent": {
                        "rng": [[4,4],[10,10]]
                    },
                    "text": {
                        "rng": [[4,4],[11,11]]
                    }
                },
                "Audios":{
                    "rng": [[3,3],[12,12]]
                }
            },
            "Noun": {
                "rng": [[2,2],[13,24]],
                1: {
                    "rng": [[3,3],[13,14]],
                    "meaning": {
                        "rng": [[4,4],[13,13]]
                    },
                    "examples": {
                        "rng": [[4,4],[14,14]]
                    }
                },
                2: {
                    "rng": [[3,3],[15,16]],
                    "meaning": {
                        "rng": [[4,4],[15,15]]
                    },
                    "examples": {
                        "rng": [[4,4],[16,16]]
                    }
                },
                "Feminine 1": {
                    "rng": [[3,3],[17,17]]
                },
                "Feminine 2": {
                    "rng": [[3,3],[18,18]]
                },
                "singular 1": {
                    "rng": [[3,3],[19,19]]
                },
                "singular 2": {
                    "rng": [[3,3],[20,20]]
                },
                "plural 1": {
                    "rng": [[3,3],[21,21]]
                },
                "plural 2": {
                    "rng": [[3,3],[22,22]]
                },
                "gender 1": {
                    "rng": [[3,3],[23,23]]
                },
                "gender 2": {
                    "rng": [[3,3],[24,24]]
                }

            },
            "Verb": {
                "rng": [[2,2],[25,28]],
                1: {
                    "rng": [[3,3],[25,26]],
                    "meaning": {
                        "rng": [[4,4],[25,25]]
                    },
                    "examples": {
                        "rng": [[4,4],[26,26]]
                    },
                },
                2: {
                    "rng": [[3,3],[27,28]],
                    "meaning": {
                        "rng": [[4,4],[27,27]]
                    },
                    "examples": {
                        "rng": [[4,4],[28,28]]
                    }
                }
            },
            "Adverb": {
                "rng": [[2,2],[29,32]],
                1: {
                    "rng": [[3,3],[29,30]],
                    "meaning": {
                        "rng": [[4,4],[29,29]]
                    },
                    "examples": {
                        "rng": [[4,4],[30,30]]
                    }
                },
                2: {
                    "rng": [[3,3],[31,32]],
                    "meaning": {
                        "rng": [[4,4],[31,31]]
                    },
                    "examples": {
                        "rng": [[4,4],[32,32]]
                    }
                }
            },
            "Adjective": {
                "rng": [[2,2],[33,44]],
                1: {
                    "rng": [[3,3],[33,34]],
                    "meaning": {
                        "rng": [[4,4],[33,33]]
                    },
                    "examples": {
                        "rng": [[4,4],[34,34]]
                    }
                },
                2: {
                    "rng": [[3,3],[35,36]],
                    "meaning": {
                        "rng": [[4,4],[35,35]]
                    },
                    "examples": {
                        "rng": [[4,4],[36,36]]
                    }
                },
                "Feminine 1": {
                    "rng": [[3,3],[37,37]]
                },
                "Feminine 2": {
                    "rng": [[3,3],[38,38]]
                },
                "common pl 1": {
                    "rng": [[3,3],[39,39]]
                },
                "common pl 2": {
                    "rng": [[3,3],[40,40]]
                },
                "mas plural 1": {
                    "rng": [[3,3],[41,41]]
                },
                "mas plural 2": {
                    "rng": [[3,3],[42,42]]
                },
                "fem plural 1": {
                    "rng": [[3,3],[43,43]]
                },
                "fem plural 2": {
                    "rng": [[3,3],[44,44]]
                }

            },
            "Pronoun": {
                "rng": [[2,2],[45,48]],
                1: {
                    "rng": [[3,3],[45,46]],
                    "meaning": {
                        "rng": [[4,4],[45,45]]
                    },
                    "examples": {
                        "rng": [[4,4],[46,46]]
                    }
                },
                2: {
                    "rng": [[3,3],[47,48]],
                    "meaning": {
                        "rng": [[4,4],[47,47]]
                    },
                    "examples": {
                        "rng": [[4,4],[48,48]]
                    }
                }
            },
            "Synonyms": {
                "rng": [[2,2],[49,49]]
            },
            "Alternative forms": {
                "rng": [[2,2],[50,50]]
            }

        },
        "Etymology 2": {
            "rng": [[1,1],[51,99]],
            "From 1": {
                "rng": [[2,2],[51,53]],
                "Language": {
                    "rng": [[3,3],[51,51]]
                },
                "Word": {
                    "rng": [[3,3],[52,52]]
                },
                "translation": {
                    "rng": [[3,3],[53,53]]
                }
            },
            "From 2": {
                "rng": [[2,2],[54,56]],
                "Language": {
                    "rng": [[3,3],[54,54]]
                },
                "Word": {
                    "rng": [[3,3],[55,55]]
                },
                "translation": {
                    "rng": [[3,3],[56,56]]
                }
            },
            "Pronunciation":{
                "rng": [[2,2],[57,61]],
                "pron_text 1": {
                    "rng": [[3,3],[57,58]],
                    "accent": {
                        "rng": [[4,4],[57,57]]
                    },
                    "text": {
                        "rng": [[4,4],[58,58]]
                    }

                },
                "pron_text 2": {
                    "rng": [[3,3],[59,60]],
                    "accent": {
                        "rng": [[4,4],[59,59]]
                    },
                    "text": {
                        "rng": [[4,4],[60,60]]
                    }
                },
                "Audios":{
                    "rng": [[3,3],[61,61]]
                }
            },
            "Noun": {
                "rng": [[2,2],[62,73]],
                1: {
                    "rng": [[3,3],[62,63]],
                    "meaning": {
                        "rng": [[4,4],[62,62]]
                    },
                    "examples": {
                        "rng": [[4,4],[63,63]]
                    }
                },
                2: {
                    "rng": [[3,3],[64,65]],
                    "meaning": {
                        "rng": [[4,4],[64,64]]
                    },
                    "examples": {
                        "rng": [[4,4],[65,65]]
                    }
                },
                "Feminine 1": {
                    "rng": [[3,3],[66,66]]
                },
                "Feminine 2": {
                    "rng": [[3,3],[67,67]]
                },
                "singular 1": {
                    "rng": [[3,3],[68,68]]
                },
                "singular 2": {
                    "rng": [[3,3],[69,69]]
                },
                "plural 1": {
                    "rng": [[3,3],[70,70]]
                },
                "plural 2": {
                    "rng": [[3,3],[71,71]]
                },
                "gender 1": {
                    "rng": [[3,3],[72,72]]
                },
                "gender 2": {
                    "rng": [[3,3],[73,73]]
                }

            },
            "Verb": {
                "rng": [[2,2],[74,77]],
                1: {
                    "rng": [[3,3],[74,75]],
                    "meaning": {
                        "rng": [[4,4],[74,74]]
                    },
                    "examples": {
                        "rng": [[4,4],[75,75]]
                    },
                },
                2: {
                    "rng": [[3,3],[76,77]],
                    "meaning": {
                        "rng": [[4,4],[76,76]]
                    },
                    "examples": {
                        "rng": [[4,4],[77,77]]
                    }
                }
            },
            "Adverb": {
                "rng": [[2,2],[78,81]],
                1: {
                    "rng": [[3,3],[78,79]],
                    "meaning": {
                        "rng": [[4,4],[78,78]]
                    },
                    "examples": {
                        "rng": [[4,4],[79,79]]
                    }
                },
                2: {
                    "rng": [[3,3],[80,81]],
                    "meaning": {
                        "rng": [[4,4],[80,80]]
                    },
                    "examples": {
                        "rng": [[4,4],[81,81]]
                    }
                }
            },
            "Adjective": {
                "rng": [[2,2],[82,93]],
                1: {
                    "rng": [[3,3],[82,83]],
                    "meaning": {
                        "rng": [[4,4],[82,82]]
                    },
                    "examples": {
                        "rng": [[4,4],[83,83]]
                    }
                },
                2: {
                    "rng": [[3,3],[84,85]],
                    "meaning": {
                        "rng": [[4,4],[84,84]]
                    },
                    "examples": {
                        "rng": [[4,4],[85,85]]
                    }
                },
                "Feminine 1": {
                    "rng": [[3,3],[86,86]]
                },
                "Feminine 2": {
                    "rng": [[3,3],[87,87]]
                },
                "common pl 1": {
                    "rng": [[3,3],[88,88]]
                },
                "common pl 2": {
                    "rng": [[3,3],[89,89]]
                },
                "mas plural 1": {
                    "rng": [[3,3],[90,90]]
                },
                "mas plural 2": {
                    "rng": [[3,3],[91,91]]
                },
                "fem plural 1": {
                    "rng": [[3,3],[92,92]]
                },
                "fem plural 2": {
                    "rng": [[3,3],[93,93]]
                }

            },
            "Pronoun": {
                "rng": [[2,2],[94,97]],
                1: {
                    "rng": [[3,3],[94,95]],
                    "meaning": {
                        "rng": [[4,4],[94,94]]
                    },
                    "examples": {
                        "rng": [[4,4],[95,95]]
                    }
                },
                2: {
                    "rng": [[3,3],[96,97]],
                    "meaning": {
                        "rng": [[4,4],[96,96]]
                    },
                    "examples": {
                        "rng": [[4,4],[97,97]]
                    }
                }
            },
            "Synonyms": {
                "rng": [[2,2],[98,98]]
            },
            "Alternative forms": {
                "rng": [[2,2],[99,99]]
            }
        }
    }

    def __init__(self, __wb, __ws):
        self.wb = __wb
        self.ws = __ws
        self.setDicKeyToExcel(self.keysDict)
        self.mergeRanges(self.keysDict)

        

    def depth(self,d):
        if isinstance(d, dict):
            return 1 + (max(map(self.depth, d.values())) if d else 0)
        return 0


    def getMergedCellVal(self,sheet, cell):
        rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
        return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value

    def getValueWithMergeLookup(self,cell):
        idx = cell.coordinate
        for range_ in self.ws.merged_cells.ranges:
            merged_cells = list(openpyxl.utils.rows_from_range(str(range_)))
            for row in merged_cells:
                if idx in row:
                    value = self.ws[merged_cells[0][0]].value
                    return value

        # return sheet.cell(idx).value
        return self.ws[idx].value

    def wordfinder2(self, searchString:str, rngr:list, rngc:list):
        rowrng = range(rngr[0], rngr[1]+1)  
        colrng = range(rngc[0], rngc[1]+1)
        print("serching for: " + searchString)
        for i in rowrng:
            for j in colrng:
                cell = self.ws.cell(row=i,column=j)
                strValue = cell.value
                if type(cell) == openpyxl.cell.cell.MergedCell:
                    # coord = cell.coordinate
                    strValue = self.getValueWithMergeLookup(cell)
                    print(strValue)

                if searchString == strValue:
                    print("found found found found found ")
                    return i,j
        return -1,-1

    def wordfinder(self, searchString:str, rngr:list, rngc:list):
        # print("serching for: " + searchString)
        for row in self.ws.iter_cols(min_row=rngr[0], min_col=rngc[0], max_row=rngr[1], max_col=rngc[1]):  
            for cell in row:  
                strValue = cell.value 
                if type(cell) == openpyxl.cell.cell.MergedCell:
                    # coord = cell.coordinate
                    strValue = self.getValueWithMergeLookup(cell)
                    # print("strVal: " + str(strValue))
                if searchString == strValue:
                    print("found found found found found ")
                    return cell.row,cell.column
        return -1,-1

    def getMergedCellRange(self, cell):
        rng = [s for s in self.ws.merged_cells.ranges if cell.coordinate in s]
        if len(rng) != 0:
            return [rng[0].min_row, rng[0].max_row], [rng[0].min_col, rng[0].max_col]

        return [cell.row, cell.row], [cell.column, cell.column]

    def cellEmpty(self, cell):
        value = cell.value
        # if type(cell) == openpyxl.cell.cell.MergedCell:
        #     return False
        if value is None or value == "":
            return True
        return False

    rangeList = [[[1,1],[2,2]]]
    # rangeDict = {}
    countEnter = 0
    countExit = 0
    count = [0,0]

    def nested_get(self, dic, keys):
        dic2 = copy.deepcopy(dic)
        for key in keys:
            dic2 = dic2[key]
        return dic2

    def nested_set(self, dic, keys, value):
        for key in keys[:-1]:
            dic = dic.setdefault(key, {})
        dic[keys[-1]] = value

    def nested_del(self, dic, keys):
        for key in keys[:-1]:
            dic = dic[key]
        del dic[keys[-1]]

    def getAllDicKeys(self, d):
        keyList = []
        for key, value in d.items():
            keyList.append(key)
            if isinstance(value, dict):
                self.getAllDicKeys(value)
        return keyList

    def setDicKeyToExcel(self, d):
        
        for key, value in d.items():
            
            if isinstance(value, dict):
                row = value["rng"][0][0]
                col = value["rng"][1][0]
                self.ws.cell(row=row, column=col).value = key
                self.setDicKeyToExcel(value)


    def updateDic(self,dic,diffrence):

        data = copy.deepcopy(dic)
        for k, v in data.copy().items():
            
            if isinstance(v, dict):     # For DICT
                data[k] = self.updateDic(v,diffrence)
            elif isinstance(v, list):   # For LIST
                # diffrence = col1 - diffrence
                tmpRng = [v[0], [v[1][0]+diffrence ,v[1][1]+diffrence]]
                data[k] = tmpRng
            # elif v == 'CHANGE ME':      # Update Key-Value
            #     # data.pop(k)
            #     # OR
            #     del data[k]
            #     data[f"{k}.CHANGED"] = 'CHANGED'
        # print(rangesDict)
        return data

    keysList = []
    def insertCell(self, data, crow, depth1):
        
        d = 0
        dp = depth1

        # it = ""
        
        for it2 in data:

            # it = it2
            
            try: 
                self.keysList[dp-1] = it2
            except:
                self.keysList.append(it2)
                
            if self.depth(data[it2]) <= d:
                keysListTmp = []
                for r in range(dp):
                    keysListTmp.append(self.keysList[r])
                # print("keys list tmp")
                # print(keysListTmp)
                # print(self.keysList)
                # tmpRng = self.nested_get(self.keysDict, self.keysList + ["rng"])
                tmpRng = self.nested_get(self.keysDict, keysListTmp + ["rng"])
                col = tmpRng[1][0]

                value = data[it2]
                if (type(value) == list):
                    value = ', '.join([str(elem) for elem in value])
                self.ws.cell(row=crow, column=col).value = value
            else:
                dp = self.insertCell(data[it2], crow, dp+1)
                # self.insertCell(data[it2], crow, dp+1)

        # self.keysList.remove(it)
        return dp - 1

    rangesDict = {"word": {"rng" : [[1,1],[2,2]]}}
    # keysList = ["word"]
    def insertCell2(self, data, crow, depth1):
        
        d = 0
        dp = depth1
        insertCol = True
        for it2 in data:
            tmpKeysList1 = []
            for i in range(dp):
                tmpKeysList1.append(self.keysList[i])
            rng = self.nested_get(self.rangesDict, tmpKeysList1+["rng"]) 
            print("searching for: " + str(it2))

            row2,col1 = self.wordfinder(it2, [dp,dp], rng[1])
            # print(tmpKeysList1)
            print("search range col: " + str(rng[1]) + " dp " + str(dp))
            col2 = col1
            
            if row2 == -1:
                col1 = rng[1][0]
                maxCol = rng[1][1]
                row2 = dp
                cCell = self.ws.cell(row=row2, column=maxCol)

                if not self.cellEmpty(cCell) or not insertCol:
                    maxCol += 1
                    self.ws.insert_cols(idx=maxCol)
                    col1 = maxCol
                col2 = maxCol

                keysListTmp = [] 
                for r in range(dp):
                    if r == 0:
                        keysListTmp.append(self.keysList[r])
                        tmpRng = self.nested_get(self.rangesDict, keysListTmp + ["rng"])
                        tmpRng[1][1] = self.ws.max_column + 1
                        self.nested_set(self.rangesDict, keysListTmp + ["rng"], tmpRng) 
                    else:
                        keysListTmp.append(self.keysList[r])
                        tmpRng = self.nested_get(self.rangesDict, keysListTmp + ["rng"])
                        tmpRng[1][1]= col2
                        self.nested_set(self.rangesDict, keysListTmp + ["rng"], tmpRng) 


                self.ws.cell(row=row2, column=col2).value = it2
            
                rngr = [dp, dp]
                rngc = [col1, col2]
                
                tmpKeysList1.append(it2)
                tmpKeysList1.append("rng")
                self.nested_set(self.rangesDict, tmpKeysList1, [rngr, rngc])
            else:
                if dp == 1:
                    it2rng = self.rangesDict['word'][it2]['rng']
                    diffrence = col1 - it2rng[1][0]
                    newData3 = self.updateDic(self.rangesDict["word"][it2], diffrence)
                    self.rangesDict["word"][it2] = newData3
                    print(self.rangesDict)

                # keysListTmp = [] 
                # for r in range(dp):
                #     keysListTmp.append(self.keysList[r])
                    # if r > dp:
                    #     print(keysListTmp)
                    #     it2rng = self.nested_get(self.rangesDict, keysListTmp + ["rng"])
                    #     diffrence = col1 - it2rng[1][0]
                    #     tmpRng = [it2rng[0], [it2rng[1][0]+diffrence ,it2rng[1][1]+diffrence]]
                    #     self.nested_set(self.rangesDict, keysListTmp + ["rng"], tmpRng)

                # if dp == 1:
                ## it2rng = self.rangesDict['word'][it2]['rng']
                # print(keysListTmp)
                # it2rng = self.nested_get(self.rangesDict, keysListTmp)
                #
                # diffrence = col1 - it2rng[it2]['rng'][1][0]
                # newData3 = self.updateDic(it2rng, diffrence)
                # # self.rangesDict["word"][it2] = newData3
                # self.nested_set(self.rangesDict, keysListTmp, newData3)
                # print(self.rangesDict)

                # it2rng = self.nested_get(self.rangesDict, tmpKeysList1+[it2,"rng"])
                # diffrence = col1 - it2rng[1][0]
                # tmpRng = [it2rng[0], [it2rng[1][0]+diffrence ,it2rng[1][1]+diffrence]]
                #
                # self.nested_set(self.rangesDict, tmpKeysList1+[it2,"rng"], tmpRng)
                # self.rangesDict["word"]["rng"] = [[1,1],[2,self.ws.max_column]]

            try: 
                self.keysList[dp] = it2
            except:
                self.keysList.append(it2)
                
            if self.depth(data[it2]) <= d:
                value = data[it2]
                if (type(value) == list):
                    value = ', '.join([str(elem) for elem in value])
                self.ws.cell(row=crow, column=col1).value = value
            else:
                self.insertCell(data[it2], crow, dp+1)

            insertCol = False

    def get_all_keys(self,d):
        for key, value in d.items():
            yield key
            if isinstance(value, dict):
                yield from self.get_all_keys(value)

    def getAllRngValues(self, d):
        for key, value in d.items():
            if key == "rng":
                yield value
            if isinstance(value, dict):
                yield from self.getAllRngValues(value)

    def mergeRanges(self, d):
        # i = 0
        print("merging ....")
        for x in self.getAllRngValues(d): 
            # if i != 0:
            # print(x)
            self.ws.merge_cells(start_row=x[0][0], start_column=x[1][0], end_row=x[0][1] , end_column=x[1][1] )
            cell = self.ws.cell(row=x[0][0], column=x[1][0])  
            # # cell.value = 'Devansh Sharma'  
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # i+=1

    def saveExcel(self, title):
        self.wb.save(title)



# wb = Workbook()  
# ws = wb.active
# dToE = DicToExcel(wb, ws)
#
# def insertData():
#     crow = 5
#
#     for it in data.keys():
#         crow+=1
#         ws.cell(row=crow, column=1).value = it
#         dToE.insertCell(data[it], crow, 1)
#
#     dToE.mergeRanges()
#     dToE.saveExcel()
#
#     # dToE.wb.save("test.xlsx")
#
# insertData()
# wb.save("test.xlsx")  


