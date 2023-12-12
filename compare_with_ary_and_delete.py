from openpyxl import Workbook, load_workbook

wb1 = load_workbook('Filtered_Darija_Terms_EN.xlsx')
ws1 = wb1.active  

wb2 = load_workbook('Darija_Terms_ary.xlsx')
ws2 = wb1.active  

wb3 = Workbook()  
ws3 = wb3.active


def termInList(term, lst):
    for word in lst:
        if term == word:
            return True
    return False


aryList = []
for i in range(1312):
    value = ws2.cell(row = i+1, column = 1).value
    aryList.append(value)

for row in ws1.iter_rows(min_row=1, max_row=1919):
    if not termInList(row[0].value, aryList):
        ws3.append((cell.value for cell in row[0:99]))


wb3.save('Final_Wt_Darija_Terms_EN.xlsx')
